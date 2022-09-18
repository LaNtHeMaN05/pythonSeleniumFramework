import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


@pytest.fixture()
def driverInit():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    return driver


@pytest.fixture()
def excelData():
    path = ".\\TestData\\Urls.xlsx"
    workbook = openpyxl.load_workbook(path)
    wb = workbook.active
    url = wb.cell(row=1, column=1).value
    username = wb.cell(row=2, column=1).value
    password = wb.cell(row=2, column=2).value
    return url,username,password

def pytest_configure(config):
    config._metadata['Project'] = 'Automation Framework'
    config._metadata['Module'] = 'Login Tests'
    config._metadata['SDET'] = 'Lan'

@pytest.mark.optionalhook
def pytest_metadata(metadata):
    metadata.pop("JAVA_HOME",None)
    metadata.pop("Plugins",None)